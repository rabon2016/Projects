#encoding=utf-8
import openpyxl
from openpyxl.styles import Border, Side, Font
import time

class ParseExcel(object):
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None)
        self.RGBDict = {'red':'FFFF3030', 'green':'FF008B00'}

    def loadWorkBook(self, excelPathAndName):
        #将Excel文件加载到内存，并获取其workbook对象
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile=excelPathAndName
        return self.workbook

    def getSheetByName(self, sheetName):
        #根据sheet名获取该sheet对象
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self, sheetIndex):
        #根据sheet的索引号获取该sheet对象
        try:
            sheetname=self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet=self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def getRowsNumber(self, sheet):
        #获取sheet中有数据区域的结束行号
        return sheet.max_row

    def getColsNumber(self, sheet):
        #获取sheet中有数据区域的结束列号
        return sheet.max_column

    def getStartRowNumber(self, sheet):
        #获取sheet中有数据区域的开始行号
        return sheet.min_row

    def getStartColNumber(self, sheet):
        #获取sheet中有数据区域的开始的列号
        return sheet.min_column

    def getRow(self, sheet, rowNo):
        #获取sheet中某一行，返回的是这一行所有的数据内容组成的tuple，下标从1开始，sheet.rows[1]表示第一行
        try:
            return list(sheet.rows)[rowNo-1]
        except Exception as e:
            raise e

    def getColumn(self, sheet, colNo):
        #获取sheet中某一列，返回的是这一列所有的数据内容组成的tuple，下标从1开始，sheet.columns[1]表示第一列
        try:
            return  list(sheet.columns)[colNo-1]
        except Exception as e:
            raise e

    def getCellOfValue(self, sheet, coordinate=None, rowNo=None, colsNo=None):
        #根据单元格所在的位置索引获取该单元格中的值，下标从1开始，
        #sheet.cell(row=1, column=1).value,表示Excel中第一行第一列的值
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def getCellOfObject(self, sheet, coordinate=None, rowNo=None, colsNo=None):
        if coordinate != None:
            try:
                return sheet.cell(coordinate = coordinate)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not  None:
            try:
                return sheet.cell(row=rowNo, column=colsNo)
            except Exception as e:
                raise e
        else:
            raise  Exception("Insufficient Coordinate of cell!")

    def writeCell(self, sheet, content, coordinate=None, rowNo=None, colsNo=None, style=None):
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = content
                if style:
                    sheet.cell(row=rowNo, column=colsNo).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinate of cell!")

    def writeCellCurrentTime(self, sheet, coordinate=None, rowNo=None, colsNo=None):
        now=int(time.time())
        timeArray = time.localtime(now)
        currenTime = time.strftime("%Y-%m-%d %H:%M:%S",timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value=currenTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = currenTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinate of cell!")

if __name__ == '__main__':
    pe=ParseExcel()
    pe.loadWorkBook(u"D:\\PyCharm 4.0.4\\DataDrivenFrameWork\\testData\\163mailContacts.xlsx")
    print("通过名称获取:",pe.getSheetByName(u"联系人").title)
    print("通过index序号获取sheet对象的名字:",pe.getSheetByIndex(0).title)
    sheet=pe.getSheetByIndex(0)
    print(type(sheet))
    print(pe.getRowsNumber(sheet))
    print(pe.getColsNumber(sheet))
    rows = pe.getRow(sheet,1)
    # print(rows)
    # for i in rowa:
    #     for j in i:
    #         print(j.value)
    # rows = pe.getRow(sheet,1)
    for i in rows:
        print(i.value)
    print(pe.getCellOfValue(sheet,rowNo=1, colsNo=1))
    pe.writeCell(sheet,u'我爱祖国',rowNo=10,colsNo=10)
    pe.writeCellCurrentTime(sheet, rowNo=10,colsNo=11)